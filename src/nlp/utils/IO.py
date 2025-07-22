import pandas as pd
import openpyxl
import os


import time
from whoosh.index import create_in, open_dir, exists_in
from whoosh.fields import Schema, TEXT, ID
from whoosh.qparser import QueryParser
from whoosh.analysis import StemmingAnalyzer



def append_to_dataset(query, sid, file_path="./dataset/会话库.xlsx", sheet_name="Sheet1", id_col="会话ID", query_col="用户问题"):

    try:
        # 读取现有数据或创建新 DataFrame
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        except FileNotFoundError:
            df = pd.DataFrame(columns=[id_col, query_col])
        
        # 创建新行数据
        new_row = {id_col: sid, query_col: query}
        # 追加到 DataFrame
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        
        # 保存回 Excel（覆盖原文件）
        df.to_excel(file_path, sheet_name=sheet_name, index=False)
        # print(f"数据已成功写入：{file_path} | 工作表：{sheet_name}")
        
    except Exception as e:
        print(f"操作失败，错误信息：{str(e)}")


def write_response_to_dataset(response):
    """将字符串写入Excel文件'大模型应答'列的末尾
    
    Args:
        response (str): 要写入的大模型回答内容
    """
    # 文件路径
    file_path = "./dataset/会话库.xlsx"
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel文件不存在: {file_path}")
    
    try:
        # 加载工作簿和工作表
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # 直接定位到标题为"大模型应答"的列（假设该列一定存在）
        target_column = next(
            cell.column for cell in sheet[1] 
            if cell.value == "大模型应答"
        )
        
        # 直接在最后一行写入数据
        last_row = sheet.max_row 
        sheet.cell(row=last_row, column=target_column).value = response
        
        # 保存文件
        workbook.save(file_path)
        return f"成功写入'大模型应答'列的第 {last_row} 行"
    
    except StopIteration:
        # 如果确实找不到列名时的处理
        raise ValueError("在标题行找不到'大模型应答'列，请检查Excel文件格式")
    except Exception as e:
        raise RuntimeError(f"写入失败: {str(e)}")
    

def get_querys_by_sid(sid: str) -> list:
    """
    从Excel文件中获取指定会话ID对应的用户问题列表
    (按行号从下往上搜索，返回顺序保持原表格的倒序)
    
    Args:
        sid: 要查找的会话ID字符串
        
    Returns:
        list: 匹配的用户问题列表，按原表格从下到上排序
    """
    # 读取Excel文件
    df = pd.read_excel('./dataset/会话库.xlsx', sheet_name='Sheet1')
    
    # 反向遍历DataFrame（从最后一行到第一行）
    result = []
    for i in range(len(df)-1, -1, -1):  # 从最后一行倒序向上遍历
        row = df.iloc[i]
        if row['会话ID'] == sid:  # 会话ID列匹配
            result.append(row['用户问题'])  # 收集用户问题列内容
    
    return result[::-1]

def get_responses_by_sid(sid: str) -> list:
    """
    从Excel文件中获取指定会话ID对应的用户问题列表
    (按行号从下往上搜索，返回顺序保持原表格的倒序)
    
    Args:
        sid: 要查找的会话ID字符串
        
    Returns:
        list: 匹配的用户问题列表，按原表格从下到上排序
    """
    # 读取Excel文件
    df = pd.read_excel('./dataset/会话库.xlsx', sheet_name='Sheet1')
    
    # 反向遍历DataFrame（从最后一行到第一行）
    result = []
    for i in range(len(df)-1, -1, -1):  # 从最后一行倒序向上遍历
        row = df.iloc[i]
        if row['会话ID'] == sid:  # 会话ID列匹配
            result.append(row['大模型应答'])  # 收集用户问题列内容
    
    return result[::-1]


def get_shopping_products():
    file_path = './dataset/购物车.xlsx'
    df = pd.read_excel(file_path)

    # 检查列名是否存在
    if '购物车' not in df.columns:
        available_columns = ', '.join(df.columns)
        raise ValueError(f"文件中不存在【购物车】列。可用的列有：{available_columns}")

    # 提取"购物车"列并转换为列表
    shopping_cart_list = df['购物车'].tolist()
    return shopping_cart_list

import os
import pandas as pd
from whoosh.index import create_in, open_dir, exists_in
from whoosh.fields import Schema, TEXT
from whoosh.qparser import QueryParser, OrGroup, FuzzyTermPlugin  
from whoosh.analysis import StemmingAnalyzer, NgramFilter
import jieba.posseg as pseg




def extract_keywords(query: str) -> str:
    """提取查询中的名词性关键词"""
    words = pseg.cut(query)
    keywords = [word for word, flag in words if flag.startswith('n')]  # 提取名词
    return " ".join(keywords) if keywords else query

def recall_alternative_products(query: str) -> list:
    """搜索产品库并返回top5结果
    
    Args:
        query: 搜索查询字符串
        
    Returns:
        top5_product: 包含top5结果的列表
    """
    # 文件路径配置
    excel_path = "./dataset/知识库.xlsx"
    index_dir = "./whoosh_index"
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel文件不存在: {excel_path}")
    
    # 读取Excel并合并三列数据
    df = pd.read_excel(excel_path)
    all_products = set()
    for col in ['一级品类', '二级品类', '三级品类']:
        if col in df.columns:
            all_products.update(df[col].dropna().astype(str).unique())
    
    # 创建Whoosh索引
    analyzer = StemmingAnalyzer() | NgramFilter(minsize=1, maxsize=5)
    schema = Schema(product=TEXT(stored=True, analyzer=analyzer))
    
    if not os.path.exists(index_dir) or not exists_in(index_dir):
        os.makedirs(index_dir, exist_ok=True)
        ix = create_in(index_dir, schema)
        
        # 添加所有产品到索引
        writer = ix.writer()
        for product in all_products:
            writer.add_document(product=product)
        writer.commit()
    else:
        ix = open_dir(index_dir)
    
    # 执行搜索（关键修改2：启用OR逻辑和模糊搜索）
    with ix.searcher() as searcher:
        processed_query = extract_keywords(query)
        
        # 关键修改3：配置OR分组查询解析器
        parser = QueryParser(
            "product", 
            ix.schema, 
            group=OrGroup  # 使用OR逻辑替代默认AND
        )
        parser.add_plugin(FuzzyTermPlugin())
        
        # 为每个关键词添加模糊符号
        fuzzy_query = " OR ".join([f"{word}~1" for word in processed_query.split()])
        q = parser.parse(fuzzy_query)
        
        results = searcher.search(q, limit=5)
        top5_product = [result['product'] for result in results]
    
    return top5_product